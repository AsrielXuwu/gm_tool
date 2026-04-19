import json
import difflib
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import openpyxl
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.cell.text import InlineFont
from openai import AzureOpenAI, NOT_GIVEN
import re
import os
import traceback
import base64
import sys
# ================= 核心配置区 =================
DEFAULT_API_VERSION = "2025-04-01-preview"

# 动态获取当前脚本/软件所在的绝对目录（兼容直接运行 .py 和打包后的 .exe）
if getattr(sys, 'frozen', False):
    APP_DIR = os.path.dirname(sys.executable)
else:
    APP_DIR = os.path.dirname(os.path.abspath(__file__))

# 强制将配置文件路径绑定到软件所在目录
CONFIG_FILE = os.path.join(APP_DIR, "lqa_api_config.enc")
SECRET_KEY = b"LQA_TOOL_SECURE_XOR_KEY_2026_!@"

def encrypt_data(data_dict):
    """字节级异或 + Base64 加密"""
    text = json.dumps(data_dict).encode('utf-8')
    encrypted = bytearray(b ^ SECRET_KEY[i % len(SECRET_KEY)] for i, b in enumerate(text))
    return base64.b64encode(encrypted).decode('utf-8')

def decrypt_data(b64_text):
    """Base64 解密 + 字节级异或还原"""
    encrypted = base64.b64decode(b64_text)
    decrypted = bytearray(b ^ SECRET_KEY[i % len(SECRET_KEY)] for i, b in enumerate(encrypted))
    return json.loads(decrypted.decode('utf-8'))
# ==============================================

ENGINES_MAP = {
    'GPT-5.3-chat': 'gpt-5.3-chat-2026-03-03',
    'GPT-5.2-chat': 'gpt-5.2-chat-2025-12-11',
    'GPT-o3-mini': 'o3-mini-2025-01-31'
}

LANGUAGES_MAP = {
    "English (United States)": "en-US",
    "English (United Kingdom)": "en-GB",
    "Chinese (Simplified, China)": "zh-CN",
    "Chinese (Traditional, Taiwan)": "zh-TW",
    "Chinese (Traditional, Hong Kong)": "zh-HK",
    "Japanese (Japan)": "ja-JP",
    "Korean (Korea)": "ko-KR",
    "Spanish (Spain)": "es-ES",
    "French (France)": "fr-FR",
    "German (Germany)": "de-DE",
    "Italian (Italy)": "it-IT",
    "Portuguese (Brazil)": "pt-BR",
    "Russian (Russia)": "ru-RU",
    "Indonesian (Indonesia)": "id-ID",
    "Thai (Thailand)": "th-TH",
    "Vietnamese (Vietnam)": "vi-VN",
    "Arabic (Saudi Arabia)": "ar-SA",
    "Polish (Poland)": "pl-PL"       # <--- 新增：波兰语
}
# ==============================================
# --- 无空格语言集合（注意：波兰语使用空格，所以千万不要加到这里面） ---
NO_SPACE_LANGS = {"zh-CN", "zh-TW", "zh-HK", "ja-JP", "th-TH"}
# ==============================================

class LQA_App:
    def __init__(self, root):
        self.root = root
        self.root.title("字幕 LQA 智能拼写检查工具")
        self.root.geometry("700x600")  # 稍微拉长窗口以容纳新组件
        
        self.file_path = tk.StringVar()
        self.output_path = tk.StringVar()
        self.stop_flag = False
        self.api_endpoint = tk.StringVar()  # 新增：绑定 UI 的 Endpoint
        self.api_key = tk.StringVar()       # 新增：绑定 UI 的 Key
        
        self.setup_ui()

    def setup_ui(self):
        # ================== 核心新增：全局可滚动面板 ==================
        # 1. 创建一个主容器铺满根窗口
        self.main_container = ttk.Frame(self.root)
        self.main_container.pack(fill="both", expand=True)

        # 2. 创建 Canvas
        self.canvas = tk.Canvas(self.main_container, highlightthickness=0)
        self.canvas.pack(side="left", fill="both", expand=True)

        # 3. 创建主垂直滚动条并绑定到 Canvas
        self.main_scrollbar = ttk.Scrollbar(self.main_container, orient="vertical", command=self.canvas.yview)
        self.main_scrollbar.pack(side="right", fill="y")
        self.canvas.configure(yscrollcommand=self.main_scrollbar.set)

        # 4. 在 Canvas 内创建一个真正的 UI 承载 Frame
        self.scrollable_frame = ttk.Frame(self.canvas)

        # 5. 将 Frame 作为窗口放入 Canvas
        self.canvas_window = self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")

        # 6. 动态更新滚动区域和宽度
        def configure_scrollregion(event):
            self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        self.scrollable_frame.bind("<Configure>", configure_scrollregion)

        def configure_canvas_width(event):
            self.canvas.itemconfig(self.canvas_window, width=event.width)
        self.canvas.bind("<Configure>", configure_canvas_width)

        # 7. 绑定鼠标全局滚轮事件 (兼容 Windows/MacOS)
        def _on_mousewheel(event):
            self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        self.canvas.bind_all("<MouseWheel>", _on_mousewheel)
        # ==============================================================

        # --- 以下所有组件的父容器从 self.root 更改为 self.scrollable_frame ---

        # --- 0. API 接口配置区 (新增) ---
        frame_api = ttk.LabelFrame(self.scrollable_frame, text="0. API 接口配置 (自动加密存储本地)", padding=10)
        frame_api.pack(fill="x", padx=10, pady=5)

        ttk.Label(frame_api, text="Azure Endpoint:").grid(row=0, column=0, sticky="w")
        ttk.Entry(frame_api, textvariable=self.api_endpoint, width=55).grid(row=0, column=1, padx=5, pady=2)

        ttk.Label(frame_api, text="API Key:").grid(row=1, column=0, sticky="w")
        ttk.Entry(frame_api, textvariable=self.api_key, width=55, show="*").grid(row=1, column=1, padx=5, pady=2)

        # 增加一个 Frame 用来纵向堆叠这两个按钮
        api_btn_frame = ttk.Frame(frame_api)
        api_btn_frame.grid(row=0, column=2, rowspan=2, padx=10, pady=2)
        
        ttk.Button(api_btn_frame, text="💾 保存配置", command=self.save_config).pack(side="top", fill="x", pady=(0, 2))
        ttk.Button(api_btn_frame, text="📂 加载配置", command=self.manual_load_config).pack(side="top", fill="x", pady=(2, 0))

        # --- 1. 文件设置区 ---
        frame_file = ttk.LabelFrame(self.scrollable_frame, text="1. 文件设置", padding=10)
        frame_file.pack(fill="x", padx=10, pady=5)
        
        ttk.Label(frame_file, text="输入 Excel:").grid(row=0, column=0, sticky="w")
        ttk.Entry(frame_file, textvariable=self.file_path, width=60).grid(row=0, column=1, padx=5, pady=2)
        ttk.Button(frame_file, text="浏览...", command=self.browse_file).grid(row=0, column=2, pady=2)

        ttk.Label(frame_file, text="输出位置(前缀):").grid(row=1, column=0, sticky="w")
        ttk.Entry(frame_file, textvariable=self.output_path, width=60).grid(row=1, column=1, padx=5, pady=2)
        ttk.Button(frame_file, text="另存为...", command=self.browse_output_file).grid(row=1, column=2, pady=2)

        # === 新增：输出模式选择 ===
        ttk.Label(frame_file, text="输出模式:").grid(row=2, column=0, sticky="w", pady=5)
        self.var_output_mode = tk.StringVar(value="split")
        mode_frame = ttk.Frame(frame_file)
        mode_frame.grid(row=2, column=1, columnspan=2, sticky="w", pady=2)
        ttk.Radiobutton(mode_frame, text="分开输出 (按Sheet和集数独立文件)", variable=self.var_output_mode, value="split").pack(side="left", padx=(0, 10))
        ttk.Radiobutton(mode_frame, text="合并输出 (全部结果保存在单一文件中)", variable=self.var_output_mode, value="merged").pack(side="left")
        # ==========================

        # --- 2. 列配置区 ---
        frame_col = ttk.LabelFrame(self.scrollable_frame, text="2. 表格列号配置 (A=1, B=2...)", padding=10)
        frame_col.pack(fill="x", padx=10, pady=5)
        
        ttk.Label(frame_col, text="原文列号:").grid(row=0, column=0, sticky="w")
        self.col_src = ttk.Entry(frame_col, width=5)
        self.col_src.insert(0, "1")
        self.col_src.grid(row=0, column=1, padx=5, sticky="w")
        
        ttk.Label(frame_col, text="译文列号:").grid(row=0, column=2, sticky="w", padx=(10, 0))
        self.col_tgt = ttk.Entry(frame_col, width=5)
        self.col_tgt.insert(0, "2")
        self.col_tgt.grid(row=0, column=3, padx=5, sticky="w")
        
        ttk.Label(frame_col, text="集数列号:").grid(row=0, column=4, sticky="w", padx=(10, 0))
        self.col_ep = ttk.Entry(frame_col, width=5)
        self.col_ep.insert(0, "3")
        self.col_ep.grid(row=0, column=5, padx=5, sticky="w")

        ttk.Label(frame_col, text="输出列号:").grid(row=1, column=0, sticky="w", pady=5)
        self.col_res = ttk.Entry(frame_col, width=5)
        self.col_res.insert(0, "4")
        self.col_res.grid(row=1, column=1, padx=5, sticky="w", pady=5)

        ttk.Label(frame_col, text="起始行号:").grid(row=1, column=2, sticky="w", padx=(10, 0), pady=5)
        self.row_start = ttk.Entry(frame_col, width=5)
        self.row_start.insert(0, "2")
        self.row_start.grid(row=1, column=3, padx=5, sticky="w", pady=5)

        self.var_with_source = tk.BooleanVar(value=True)
        ttk.Checkbutton(frame_col, text="双语模式 (参考原文)", variable=self.var_with_source).grid(row=1, column=4, columnspan=2, sticky="w", padx=(10, 0))

        # --- 3. 工作表 (Sheet) 过滤 ---
        frame_sheets = ttk.LabelFrame(self.scrollable_frame, text="3. 工作表 (Sheet) 选择 (不选默认仅检查第1个Sheet)", padding=10)
        frame_sheets.pack(fill="x", padx=10, pady=5)
        
        self.btn_load_sheets = ttk.Button(frame_sheets, text="📑 扫描 Sheet", command=self.load_sheets)
        self.btn_load_sheets.pack(side="left", padx=5)

        self.sheet_listbox = tk.Listbox(frame_sheets, selectmode=tk.MULTIPLE, height=3, exportselection=False)
        self.sheet_listbox.pack(side="left", fill="both", expand=True, padx=5)
        
        sheet_scrollbar = ttk.Scrollbar(frame_sheets, orient="vertical", command=self.sheet_listbox.yview)
        sheet_scrollbar.pack(side="left", fill="y")
        self.sheet_listbox.config(yscrollcommand=sheet_scrollbar.set)

        # --- 4. 集数过滤区 ---
        frame_eps = ttk.LabelFrame(self.scrollable_frame, text="4. 集数过滤 (支持 Ctrl 多选；不选默认检查全部)", padding=10)
        frame_eps.pack(fill="x", padx=10, pady=5)
        
        self.btn_load_eps = ttk.Button(frame_eps, text="🔍 扫描集数", command=self.load_episodes)
        self.btn_load_eps.pack(side="left", padx=5)

        self.ep_listbox = tk.Listbox(frame_eps, selectmode=tk.MULTIPLE, height=4, exportselection=False)
        self.ep_listbox.pack(side="left", fill="both", expand=True, padx=5)

        scrollbar = ttk.Scrollbar(frame_eps, orient="vertical", command=self.ep_listbox.yview)
        scrollbar.pack(side="left", fill="y")
        self.ep_listbox.config(yscrollcommand=scrollbar.set)

        # --- 5. AI 参数配置区 ---
        frame_ai = ttk.LabelFrame(self.scrollable_frame, text="5. AI 引擎及要求设置", padding=10)
        frame_ai.pack(fill="x", padx=10, pady=5)
        
        ttk.Label(frame_ai, text="选择模型:").grid(row=0, column=0, sticky="w")
        self.model_box = ttk.Combobox(frame_ai, values=list(ENGINES_MAP.keys()), width=20)
        self.model_box.current(0)
        self.model_box.grid(row=0, column=1, padx=5, sticky="w")

        ttk.Label(frame_ai, text="目标语言:").grid(row=0, column=2, sticky="w", padx=(10,0))
        self.lang_box = ttk.Combobox(frame_ai, values=list(LANGUAGES_MAP.keys()), width=25)
        self.lang_box.set("English (United States)")
        self.lang_box.grid(row=0, column=3, padx=5, sticky="w")

        ttk.Label(frame_ai, text="Token 上限/次:").grid(row=1, column=0, sticky="w", pady=5)
        self.token_limit = ttk.Entry(frame_ai, width=10)
        self.token_limit.insert(0, "2000")
        self.token_limit.grid(row=1, column=1, padx=5, sticky="w", pady=5)

        ttk.Label(frame_ai, text="要求/背景:").grid(row=2, column=0, sticky="nw", pady=5)
        self.context_text = tk.Text(frame_ai, width=65, height=3)
        self.context_text.insert("1.0", "这通常是短剧的字幕文件，背景是古代仙侠背景。请确保称呼和语气符合设定，修改错别字、不地道的表达和拼写错误，保留专有名词。")
        self.context_text.grid(row=2, column=1, columnspan=3, padx=5, pady=5)

        # --- 6. 操作与日志区 ---
        frame_action = tk.Frame(self.scrollable_frame)
        frame_action.pack(fill="both", expand=True, padx=10, pady=5)

        # === 新增：实时统计信息 ===
        self.var_stats = tk.StringVar(value="准备就绪")
        lbl_stats = ttk.Label(frame_action, textvariable=self.var_stats, font=("Arial", 10, "bold"))
        lbl_stats.pack(pady=(5, 0))
        # ==========================

        btn_container = tk.Frame(frame_action)
        btn_container.pack(pady=5)
        
        self.btn_start = ttk.Button(btn_container, text="🚀 开始检查", command=self.start_processing)
        self.btn_start.pack(side="left", padx=10)

        self.btn_stop = ttk.Button(btn_container, text="🛑 停止检查", command=self.stop_processing, state="disabled")
        self.btn_stop.pack(side="left", padx=10)

        # 将 height=13 修改为你想要的高度，比如 20 或 25
        self.log_area = scrolledtext.ScrolledText(frame_action, width=85, height=30, state='disabled')
        self.log_area.pack(fill="both", expand=True)
        self.load_config()
    
    def save_config(self):
        """加密并保存 API 配置到软件所在目录"""
        data = {
            "endpoint": self.api_endpoint.get().strip(),
            "api_key": self.api_key.get().strip()
        }
        if not data["endpoint"] or not data["api_key"]:
            messagebox.showwarning("警告", "Endpoint 和 API Key 不能为空！")
            return
            
        try:
            enc_data = encrypt_data(data)
            # 使用绝对路径 CONFIG_FILE 强制保存到软件目录
            with open(CONFIG_FILE, "w", encoding="utf-8") as f:
                f.write(enc_data)
            messagebox.showinfo("成功", f"API 配置已加密保存至:\n{CONFIG_FILE}")
        except Exception as e:
            messagebox.showerror("错误", f"保存配置失败 (请检查目录是否有写入权限):\n{e}")

    def manual_load_config(self):
        """手动弹出对话框选择配置文件并加载"""
        filepath = filedialog.askopenfilename(
            title="选择 API 配置文件",
            filetypes=[("Encrypted Config", "*.enc"), ("All Files", "*.*")],
            initialdir=APP_DIR  # 默认打开软件所在目录
        )
        if filepath:
            self.load_config(filepath)

    def load_config(self, filepath=CONFIG_FILE):
        """启动时默认从软件所在目录加载，或从指定路径加载"""
        if os.path.exists(filepath):
            try:
                with open(filepath, "r", encoding="utf-8") as f:
                    enc_data = f.read()
                data = decrypt_data(enc_data)
                self.api_endpoint.set(data.get("endpoint", ""))
                self.api_key.set(data.get("api_key", ""))
                
                # 如果是手动加载的，给个成功提示
                if filepath != CONFIG_FILE:
                    messagebox.showinfo("成功", "配置文件加载成功！")
            except Exception as e:
                # 只有手动加载报错时才弹窗，启动自动加载报错只写日志
                if filepath != CONFIG_FILE:
                    messagebox.showerror("错误", f"配置文件解析失败或已损坏:\n{e}")
                else:
                    print(f"自动加载配置失败: {e}")

    def browse_file(self):
        filepath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if filepath:
            self.file_path.set(filepath)
            default_output = filepath.replace(".xlsx", "_CheckedResult.xlsx")
            self.output_path.set(default_output)

    def browse_output_file(self):
        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="选择保存前缀位置"
        )
        if filepath:
            self.output_path.set(filepath)

    def log(self, message):
        self.root.after(0, self._append_log, message)

    def _append_log(self, message):
        self.log_area.config(state='normal')
        self.log_area.insert(tk.END, message + "\n")
        self.log_area.see(tk.END)
        self.log_area.config(state='disabled')

    def load_sheets(self):
        """扫描 Excel 中的所有 Sheet 并加载到列表"""
        input_file = self.file_path.get()
        if not input_file:
            messagebox.showerror("错误", "请先在上方选择输入 Excel 文件！")
            return
        try:
            self.log("正在扫描工作表 (Sheet)...")
            self.root.update()
            
            wb = openpyxl.load_workbook(input_file, read_only=True)
            sheets = wb.sheetnames
            wb.close()
            
            self.sheet_listbox.delete(0, tk.END)
            for s in sheets:
                self.sheet_listbox.insert(tk.END, s)
                
            self.log(f"✅ 成功扫描到 {len(sheets)} 个工作表。")
        except Exception as e:
            self.log(f"❌ 扫描工作表失败: {e}")
            messagebox.showerror("错误", f"读取工作表失败：\n{e}")

    def load_episodes(self):
        """扫描选中的 Sheet 中的集数（带去重与自然排序）"""
        input_file = self.file_path.get()
        if not input_file:
            messagebox.showerror("错误", "请先在上方选择输入 Excel 文件！")
            return
            
        try:
            c_ep = int(self.col_ep.get())
            r_start = int(self.row_start.get())
            
            self.log("正在扫描集数，请稍候...")
            self.root.update()
            
            wb = openpyxl.load_workbook(input_file, data_only=True)
            
            # --- 核心：只在选中的 Sheet 中扫描集数 ---
            selected_sheet_indices = self.sheet_listbox.curselection()
            if selected_sheet_indices:
                target_sheets = [self.sheet_listbox.get(i) for i in selected_sheet_indices]
            else:
                target_sheets = [wb.sheetnames[0]] # 默认第一个
            
            episodes = set()
            for sheet_name in target_sheets:
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    for row in range(r_start, ws.max_row + 1):
                        ep_val = str(ws.cell(row=row, column=c_ep).value or "未分类集数").strip()
                        if ep_val:
                            episodes.add(ep_val)
            wb.close()
            
            def natural_sort_key(s):
                return [int(text) if text.isdigit() else text.lower() for text in re.split('([0-9]+)', s)]
                
            sorted_eps = sorted(list(episodes), key=natural_sort_key)
            
            self.ep_listbox.delete(0, tk.END)
            for ep in sorted_eps:
                self.ep_listbox.insert(tk.END, ep)
                
            self.log(f"✅ 成功从目标 Sheet 中扫描了 {len(sorted_eps)} 个独立集数。")
            
        except Exception as e:
            self.log(f"❌ 扫描集数失败: {e}")
            messagebox.showerror("错误", f"读取集数失败，请检查格式或列号是否正确：\n{e}")

    def start_processing(self):
        if not self.file_path.get() or not self.output_path.get():
            messagebox.showerror("错误", "请先选择输入文件并设置输出路径！")
            return
        
        self.stop_flag = False
        self.total_tokens_used = 0 
        self.var_stats.set("正在初始化任务...") # <--- 新增状态更新
        self.btn_start.config(state="disabled")
        self.btn_load_eps.config(state="disabled")
        self.btn_load_sheets.config(state="disabled")
        self.btn_stop.config(state="normal")
        self.log_area.config(state='normal')
        self.log_area.delete(1.0, tk.END)
        self.log_area.config(state='disabled')
        
        threading.Thread(target=self.process_excel_worker, daemon=True).start()

    def stop_processing(self):
        self.stop_flag = True
        self.btn_stop.config(state="disabled")
        self.log("⚠️ 收到停止指令！当前批次请求完成后将安全退出...")

    def estimate_tokens(self, text):
        try:
            import tiktoken
            encoding = tiktoken.get_encoding('cl100k_base')
            return len(encoding.encode(text))
        except ImportError:
            return int(len(str(text)) * 1.5)

    def build_prompt(self, target_lang_name, target_lang_code, additional_context, with_src):
        if target_lang_code in NO_SPACE_LANGS:
            concat_rule = f"Since {target_lang_name} does not use spaces between words, if a sentence spans across multiple lines (IDs), treat them as directly connected without any spaces when evaluating context and grammar."
        else:
            concat_rule = f"Since {target_lang_name} uses spaces between words, if a sentence spans across multiple lines (IDs), treat them as connected with a space when evaluating context and grammar."

        if with_src:
            sys_prompt = f"""# System Role:
You are a subtitle LQA expert.
# Context:
- Target Language: {target_lang_name} ({target_lang_code})
- Continues Rule: {concat_rule}
- Reqs: {additional_context}
# JSON Fields:
i: id
s: Source
t: Translation
r: revisedTranslation
# Format:
In: [{{\"i\":\"1\",\"s\":\"Hi\",\"t\":\"哈喽\"}}]
Out: {{\"result\":[{{\"i\":\"1\",\"r\":\"你好\"}}]}}
# Task:
Fix spelling/grammar in `t` using `s` as context. Return ONLY valid minified JSON.
"""
        else:
            sys_prompt = f"""# System Role:
You are a proofreading expert.
# Context:
- Target Language: {target_lang_name} ({target_lang_code})
- Continues Rule: {concat_rule}
- Reqs: {additional_context}
# JSON Fields:
i: id
t: Translation
r: revisedTranslation
# Format:
In: [{{\"i\":\"1\",\"t\":\"哈喽\"}}]
Out: {{\"result\":[{{\"i\":\"1\",\"r\":\"你好\"}}]}}
# Task:
Fix spelling/grammar in `t`. Return ONLY valid minified JSON.
"""
        return sys_prompt
    
    def get_rich_text_diff(self, old_text, new_text):
        # 如果没有变动，直接返回两个纯文本
        if old_text == new_text:
            return old_text, new_text

        blue_font = InlineFont(color="0000FF") # 原始译文的变动用蓝色
        red_font = InlineFont(color="FF0000")  # 修改后的变动用红色
        
        rich_old = CellRichText()
        rich_new = CellRichText()
        
        matcher = difflib.SequenceMatcher(None, old_text, new_text)
        for tag, i1, i2, j1, j2 in matcher.get_opcodes():
            if tag == 'equal':
                rich_old.append(old_text[i1:i2])
                rich_new.append(new_text[j1:j2])
            elif tag == 'insert':
                # 新增的词：在旧文本里不存在，在新文本里标红
                rich_new.append(TextBlock(font=red_font, text=new_text[j1:j2]))
            elif tag == 'delete':
                # 删除的词：在旧文本里标蓝，在新文本里直接忽略(取消下划线)
                rich_old.append(TextBlock(font=blue_font, text=old_text[i1:i2]))
            elif tag == 'replace':
                # 替换的词：在旧文本里标蓝，在新文本里标红
                rich_old.append(TextBlock(font=blue_font, text=old_text[i1:i2]))
                rich_new.append(TextBlock(font=red_font, text=new_text[j1:j2]))
                
        return rich_old, rich_new
    
    def process_excel_worker(self):
        try:
            import time
            input_file = self.file_path.get()
            output_file_base = self.output_path.get()
            
            c_src = int(self.col_src.get())
            c_tgt = int(self.col_tgt.get())
            c_ep = int(self.col_ep.get())
            c_res = int(self.col_res.get())
            r_start = int(self.row_start.get())
            
            with_src = self.var_with_source.get()
            t_limit = int(self.token_limit.get())
            additional_context = self.context_text.get("1.0", tk.END).strip()
            output_mode = self.var_output_mode.get()

            ui_lang_name = self.lang_box.get()
            target_lang_code = LANGUAGES_MAP.get(ui_lang_name, ui_lang_name)
            
            ui_model_name = self.model_box.get()
            actual_deployment_name = ENGINES_MAP.get(ui_model_name, ui_model_name)

            selected_indices = self.ep_listbox.curselection()
            selected_episodes = [self.ep_listbox.get(i) for i in selected_indices]
            
            selected_sheet_indices = self.sheet_listbox.curselection()
            
            self.log(f"正在加载原始 Excel 文件: {input_file}")
            wb = openpyxl.load_workbook(input_file)
            
            if selected_sheet_indices:
                target_sheets = [self.sheet_listbox.get(i) for i in selected_sheet_indices]
            else:
                target_sheets = [wb.sheetnames[0]]
                
            # === 拦截校验：检查是否配置了 API ===
            api_endpoint_val = self.api_endpoint.get().strip()
            api_key_val = self.api_key.get().strip()
            
            if not api_endpoint_val or not api_key_val:
                self.log("❌ 错误：请求被拦截。请先在界面最上方配置并保存 API 接口信息！")
                self.root.after(0, lambda: messagebox.showerror("错误", "API Endpoint 和 Key 不能为空，请先配置！"))
                return

            client = AzureOpenAI(
                azure_endpoint=api_endpoint_val,
                api_key=api_key_val,
                api_version=DEFAULT_API_VERSION
            )

            sys_prompt = self.build_prompt(ui_lang_name, target_lang_code, additional_context, with_src)

            # === 1. 预扫描：构建任务池并计算总集数 ===
            all_tasks = []
            
            if output_mode == "merged":
                for s_name in target_sheets:
                    if s_name in wb.sheetnames:
                        wb[s_name].cell(row=r_start-1, column=c_res, value="Spell Check Result (AI)")

            for sheet_name in target_sheets:
                if sheet_name not in wb.sheetnames:
                    continue
                ws = wb[sheet_name]
                
                headers = []
                for r in range(1, r_start):
                    headers.append([ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)])

                episodes_data = {}
                for row in range(r_start, ws.max_row + 1):
                    tgt_text = ws.cell(row=row, column=c_tgt).value
                    if not tgt_text:
                        continue
                    
                    ep_val = str(ws.cell(row=row, column=c_ep).value or "未分类集数").strip()
                    original_row_values = [ws.cell(row=row, column=c).value for c in range(1, ws.max_column + 1)]
                    
                    item = {"i": str(row), "t": str(tgt_text)}
                    if with_src:
                        src_text = ws.cell(row=row, column=c_src).value
                        item["s"] = str(src_text) if src_text else ""
                    
                    if ep_val not in episodes_data:
                        episodes_data[ep_val] = []
                    
                    episodes_data[ep_val].append({
                        "id": str(row),
                        "item": item,
                        "original_row_values": original_row_values
                    })
                    
                for ep_name, rows_in_ep in episodes_data.items():
                    if selected_episodes and ep_name not in selected_episodes:
                        continue
                    all_tasks.append({
                        "sheet_name": sheet_name,
                        "ep_name": ep_name,
                        "rows_in_ep": rows_in_ep,
                        "headers": headers,
                        "ws": ws
                    })

            total_eps = len(all_tasks)
            success_count = 0
            fail_count = 0
            error_logs = []
            
            if total_eps == 0:
                self.log("⚠️ 没有找到需要处理的集数！")
                self.root.after(0, lambda: self.btn_start.config(state="normal"))
                return
            
            # === 2. 逐集处理，带断点容错 ===
            for current_idx, task in enumerate(all_tasks, start=1):
                if self.stop_flag:
                    break
                    
                sheet_name = task["sheet_name"]
                ep_name = task["ep_name"]
                rows_in_ep = task["rows_in_ep"]
                headers = task["headers"]
                ws = task["ws"]

                self.root.after(0, lambda c=current_idx, t=total_eps, s=success_count, f=fail_count: 
                                self.var_stats.set(f"正在处理第{c}集，共{t}集，成功{s}集，失败{f}集"))
                self.log(f"\n====== 处理中: Sheet [{sheet_name}] -> 集数 [{ep_name}] (共 {len(rows_in_ep)} 行) ======")

                current_batch = []
                current_tokens = 0
                episode_results = []
                has_error = False
                error_msg = ""

                # 发送请求 (不写入)
                for idx, item_data in enumerate(rows_in_ep):
                    row_item = item_data["item"]
                    item_json_str = json.dumps(row_item, ensure_ascii=False)
                    item_tokens = self.estimate_tokens(item_json_str)
                    
                    if current_tokens + item_tokens > t_limit and current_batch:
                        try:
                            res = self._send_batch_request(client, ui_model_name, actual_deployment_name, sys_prompt, current_batch)
                            episode_results.extend(res)
                        except Exception as e:
                            has_error = True
                            error_msg = str(e)
                            break  # 立即中止当前集的检查
                        current_batch = []
                        current_tokens = 0
                        
                    current_batch.append(row_item)
                    current_tokens += item_tokens
                    
                    if idx == len(rows_in_ep) - 1 and current_batch:
                        try:
                            res = self._send_batch_request(client, ui_model_name, actual_deployment_name, sys_prompt, current_batch)
                            episode_results.extend(res)
                        except Exception as e:
                            has_error = True
                            error_msg = str(e)
                            break

                # 失败拦截：如果报错，跳过 Excel 写入阶段并记录
                if has_error:
                    fail_count += 1
                    error_logs.append([os.path.basename(input_file), f"[{sheet_name}] {ep_name}", error_msg])
                    self.log(f"  ❌ 【检查失败拦截】本集已跳过写入。错误原因: {error_msg}")
                    continue

                # ================= 如果全部成功，统一处理 Excel 写入 =================
                success_count += 1
                self.root.after(0, lambda c=current_idx, t=total_eps, s=success_count, f=fail_count: 
                                self.var_stats.set(f"正在处理第{c}集，共{t}集，成功{s}集，失败{f}集"))

                # 建立 AI 返回数据的映射字典
                res_mapping = {str(r.get("i", r.get("id"))): r.get("r", r.get("revisedTranslation", "")) for r in episode_results}

                try:
                    if output_mode == "split":
                        ep_wb = openpyxl.Workbook()
                        ep_ws = ep_wb.active
                        ep_ws.title = "LQA Result"
                        
                        for h_row_idx, h_row_vals in enumerate(headers, start=1):
                            for col_idx, val in enumerate(h_row_vals, start=1):
                                ep_ws.cell(row=h_row_idx, column=col_idx, value=val)
                            ep_ws.cell(row=h_row_idx, column=c_res, value="Spell Check Result (AI)")

                        current_new_row_idx = r_start
                        for item_data in rows_in_ep:
                            for col_idx, val in enumerate(item_data["original_row_values"], start=1):
                                ep_ws.cell(row=current_new_row_idx, column=col_idx, value=val)

                            row_id = item_data["id"]
                            old_text = item_data["item"].get("t", item_data["item"].get("Translation", ""))
                            new_text = res_mapping.get(row_id, old_text) # 匹配不到则沿用旧文
                            
                            # --- 核心修改：获取双向对比的富文本 ---
                            rich_old, rich_new = self.get_rich_text_diff(old_text, new_text)
                            
                            # 覆盖原来的译文列 (c_tgt)，变动处标蓝
                            ep_ws.cell(row=current_new_row_idx, column=c_tgt).value = rich_old
                            # 写入新的结果列 (c_res)，变动处标红
                            ep_ws.cell(row=current_new_row_idx, column=c_res).value = rich_new
                            # ------------------------------------
                            
                            current_new_row_idx += 1
                        
                        safe_sheet_name = str(sheet_name).replace("/", "_").replace("\\", "_")
                        safe_ep_name = str(ep_name).replace("/", "_").replace("\\", "_")
                        out_dir = os.path.dirname(output_file_base)
                        out_name = os.path.basename(output_file_base).replace(".xlsx", "")
                        ep_output_file = os.path.join(out_dir, f"{out_name}_{safe_sheet_name}_{safe_ep_name}.xlsx")
                        ep_wb.save(ep_output_file)
                        self.log(f"  💾 【分发保存】{sheet_name} - {ep_name} 已成功保存。")
                    else:
                        for item_data in rows_in_ep:
                            row_id = item_data["id"]
                            mapping_row = int(row_id)
                            old_text = item_data["item"].get("t", item_data["item"].get("Translation", ""))
                            new_text = res_mapping.get(row_id, old_text)
                            
                            # --- 核心修改：获取双向对比的富文本 ---
                            rich_old, rich_new = self.get_rich_text_diff(old_text, new_text)
                            
                            # 覆盖原来的译文列 (c_tgt)，变动处标蓝
                            ws.cell(row=mapping_row, column=c_tgt).value = rich_old
                            # 写入新的结果列 (c_res)，变动处标红
                            ws.cell(row=mapping_row, column=c_res).value = rich_new
                            # ------------------------------------
                            
                        wb.save(output_file_base)
                        self.log(f"  💾 【合并进度追加】{sheet_name} - {ep_name} 已安全追加至原文件。")
                except Exception as e:
                    self.log(f"  ⚠️ 写入或保存时出错: {e}")

            # === 3. 生成错误报告 ===
            if error_logs:
                try:
                    err_wb = openpyxl.Workbook()
                    err_ws = err_wb.active
                    err_ws.append(["文件名", "集数", "错误信息"])
                    for log_entry in error_logs:
                        err_ws.append(log_entry)
                    
                    out_dir = os.path.dirname(output_file_base)
                    err_file = os.path.join(out_dir, f"LQA_Error_Report_{int(time.time())}.xlsx")
                    err_wb.save(err_file)
                    self.log(f"\n⚠️ 发现 {fail_count} 个请求失败的集数，错误报告已生成: {err_file}")
                except Exception as e:
                    self.log(f"\n⚠️ 错误报告生成失败: {e}")

            # ================= 循环处理结束提示 =================
            if self.stop_flag:
                self.log(f"\n====== 🛑 检查已被手动终止 ======")
                self.log(f"💰 本次共计消耗 Token: {self.total_tokens_used}")
                self.root.after(0, lambda: messagebox.showinfo("已终止", f"检查任务已终止。\n共消耗 Token: {self.total_tokens_used}\n成功: {success_count} 集, 失败: {fail_count} 集。"))
            else:
                self.log(f"\n====== ✨ 设定任务处理完成！ ======")
                self.log(f"💰 本次共计消耗 Token: {self.total_tokens_used}")
                self.root.after(0, lambda: messagebox.showinfo("完成", f"所选任务处理完成！\n共消耗 Token: {self.total_tokens_used}\n成功: {success_count} 集, 失败: {fail_count} 集。"))

        except Exception as e:
            self.log(f"❌ 发生致命错误: {str(e)}")
            import traceback
            traceback.print_exc()
        finally:
            self.root.after(0, lambda: self.btn_start.config(state="normal"))
            self.root.after(0, lambda: self.btn_load_eps.config(state="normal"))
            self.root.after(0, lambda: self.btn_load_sheets.config(state="normal"))
            self.root.after(0, lambda: self.btn_stop.config(state="disabled"))

    def _send_batch_request(self, client, ui_model_name, actual_deployment_name, sys_prompt, batch_data):
        # 抛出异常由外层捕获拦截
        start_id = batch_data[0].get('i', batch_data[0].get('id'))
        end_id = batch_data[-1].get('i', batch_data[-1].get('id'))
        self.log(f"  -> 发送请求... [引擎: {ui_model_name}] (原表行号: {start_id} 至 {end_id})")
        
        user_prompt = f"# Work Data:\n{json.dumps(batch_data, ensure_ascii=False)}"
        
        target_temperature = 0.3
        json_resp = True
        
        if '4o' not in ui_model_name and '4.1' not in ui_model_name:
            target_temperature = 1.0
            
        if 'o1' in ui_model_name or 'o3' in ui_model_name or 'o4' in ui_model_name:
            json_resp = False

        request_kwargs = {
            "model": actual_deployment_name,
            "messages": [
                {"role": "system", "content": sys_prompt},
                {"role": "user", "content": user_prompt}
            ],
            "temperature": target_temperature
        }
        
        if json_resp:
            request_kwargs["response_format"] = {"type": "json_object"}
        
        response = client.chat.completions.create(**request_kwargs)
        
        if hasattr(response, 'usage') and response.usage:
            used_tokens = response.usage.total_tokens
            self.total_tokens_used += used_tokens
            self.log(f"  [计量] 本次请求花费: {used_tokens} tokens | 累计花费: {self.total_tokens_used} tokens")
            
        resp_content = response.choices[0].message.content
        
        clean_content = resp_content.strip()
        if clean_content.startswith("```json"):
            clean_content = clean_content[7:]
        elif clean_content.startswith("```"):
            clean_content = clean_content[3:]
        if clean_content.endswith("```"):
            clean_content = clean_content[:-3]
            
        result_data = json.loads(clean_content.strip()).get("result", [])
        self.log(f"  ✅ 成功接收 {len(result_data)} 行数据。")
        
        return result_data
    
if __name__ == "__main__":
    root = tk.Tk()
    app = LQA_App(root)
    root.mainloop()